"""Unit tests for the aiogram-independent .xlsx report builder (Queue 2.2).

Verifies the generated workbook is valid (opens with openpyxl), carries the
expected header/blocks/values, and that money cells are real numbers with a
number_format — not strings."""
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.services.calc.export import (
    REPORT_TITLE,
    SITE_URL,
    CalcExportData,
    build_calc_workbook,
    export_filename,
)


def _sample(**overrides) -> CalcExportData:
    data = dict(
        run_id=42,
        device_name="Antminer S19 Pro",
        quantity=3,
        currency="USDT",
        created_at=datetime(2026, 6, 20, 14, 30, tzinfo=timezone.utc),
        hashrate_ths=Decimal("110.00"),
        power_w=3250,
        power_price=Decimal("0.0500"),
        net_profit_day=Decimal("12.3456"),
        net_profit_month=Decimal("370.37"),
    )
    data.update(overrides)
    return CalcExportData(**data)


def _load(payload: bytes):
    return load_workbook(BytesIO(payload))


def _all_text(ws) -> str:
    return "\n".join(
        str(c.value)
        for row in ws.iter_rows()
        for c in row
        if c.value is not None
    )


def test_workbook_is_valid_and_has_title_and_footer():
    ws = _load(build_calc_workbook(_sample())).active
    text = _all_text(ws)
    assert REPORT_TITLE in text
    assert SITE_URL in text
    assert "Параметры" in text
    assert "Результаты" in text


def test_workbook_contains_parameter_values():
    ws = _load(build_calc_workbook(_sample())).active
    text = _all_text(ws)
    assert "Antminer S19 Pro" in text
    assert "Дата расчёта" in text
    assert "20.06.2026" in text


def _find_value_cell(ws, label_fragment: str):
    """Return the column-B value cell for the row whose A label matches."""
    for row in ws.iter_rows():
        label = row[0].value
        if label and label_fragment in str(label):
            return row[1]
    return None


def test_money_cells_are_numbers_with_format():
    ws = _load(build_calc_workbook(_sample())).active

    day = _find_value_cell(ws, "прибыль в день")
    assert day is not None
    assert isinstance(day.value, (int, float))
    assert abs(day.value - 12.3456) < 1e-6
    assert day.number_format and "0" in day.number_format

    month = _find_value_cell(ws, "прибыль в месяц")
    assert isinstance(month.value, (int, float))
    assert abs(month.value - 370.37) < 1e-6


def test_numeric_param_cells_are_numbers():
    ws = _load(build_calc_workbook(_sample())).active
    qty = _find_value_cell(ws, "Количество")
    assert isinstance(qty.value, (int, float))
    assert qty.value == 3
    price = _find_value_cell(ws, "Цена э/э")
    assert isinstance(price.value, (int, float))
    assert abs(price.value - 0.05) < 1e-9


def test_manual_entry_without_specs_still_valid():
    # Manual calc may omit model id, but device_name + headline always present.
    data = _sample(device_name="Своё оборудование")
    ws = _load(build_calc_workbook(data)).active
    assert "Своё оборудование" in _all_text(ws)


def test_missing_optional_values_render_dash():
    data = _sample(net_profit_month=None, power_price=None, hashrate_ths=None)
    ws = _load(build_calc_workbook(data)).active
    month = _find_value_cell(ws, "прибыль в месяц")
    assert month.value == "—"


def test_currency_symbol_used_in_labels():
    ws = _load(build_calc_workbook(_sample(currency="RUB"))).active
    text = _all_text(ws)
    assert "₽" in text


def test_export_filename_shape():
    name = export_filename(
        7, datetime(2026, 1, 5, tzinfo=timezone.utc)
    )
    assert name == "kirkalab_calc_7_2026-01-05.xlsx"


def test_export_filename_defaults_to_today_when_no_date():
    name = export_filename(9, None)
    assert name.startswith("kirkalab_calc_9_")
    assert name.endswith(".xlsx")
