"""Integration tests for the Excel export endpoint (Queue 2.2).

Covers the PRO gate (FREE -> 403, PRO -> file), ownership scoping (a run that
belongs to another user / does not exist -> 404), and that the returned bytes
are a valid .xlsx with the expected content type + filename."""
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.crud import calc as crud_calc
from app.crud import users as crud_users
from app.db import models

BOT_SECRET = "test-bot-secret"

XLSX_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _headers():
    return {"X-Bot-Secret": BOT_SECRET}


def _make_user(db, telegram_user_id: int, *, is_pro: bool) -> models.User:
    user = crud_users.get_or_create_telegram_user(
        db, telegram_user_id=telegram_user_id
    )
    if is_pro:
        user.is_pro = True
        user.premium_until = datetime(2099, 1, 1, tzinfo=timezone.utc)
        db.add(user)
        db.commit()
        db.refresh(user)
    return user


def _make_run(db, user_id: int) -> models.CalculationRun:
    return crud_calc.record_run(
        db,
        user_id=user_id,
        device_model_id=None,
        device_name="Antminer S19 Pro",
        hashrate_ths=Decimal("110.00"),
        power_w=3250,
        quantity=2,
        power_price=Decimal("0.0500"),
        currency="USDT",
        net_profit_day_usdt=Decimal("12.34"),
        net_profit_month_usdt=Decimal("370.20"),
    )


def test_export_requires_bot_secret(client, db):
    user = _make_user(db, 880001, is_pro=True)
    run = _make_run(db, user.id)
    resp = client.get(
        f"/api/v1/internal/calc/{run.id}/export.xlsx",
        params={"telegram_user_id": 880001},
    )
    assert resp.status_code == 403


def test_free_user_gets_403_upsell(client, db):
    user = _make_user(db, 880002, is_pro=False)
    run = _make_run(db, user.id)
    resp = client.get(
        f"/api/v1/internal/calc/{run.id}/export.xlsx",
        params={"telegram_user_id": 880002},
        headers=_headers(),
    )
    assert resp.status_code == 403, resp.text
    assert "PRO" in resp.json()["detail"]


def test_pro_user_gets_valid_xlsx(client, db):
    user = _make_user(db, 880003, is_pro=True)
    run = _make_run(db, user.id)
    resp = client.get(
        f"/api/v1/internal/calc/{run.id}/export.xlsx",
        params={"telegram_user_id": 880003},
        headers=_headers(),
    )
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(XLSX_TYPE)
    disposition = resp.headers["content-disposition"]
    assert f"kirkalab_calc_{run.id}_" in disposition
    assert disposition.endswith(".xlsx") or ".xlsx" in disposition

    wb = load_workbook(BytesIO(resp.content))
    text = "\n".join(
        str(c.value)
        for row in wb.active.iter_rows()
        for c in row
        if c.value is not None
    )
    assert "Antminer S19 Pro" in text
    assert "kirkalab.ru" in text


def test_cannot_export_other_users_run(client, db):
    owner = _make_user(db, 880004, is_pro=True)
    run = _make_run(db, owner.id)
    # A *different* PRO user must not be able to export the owner's run.
    _make_user(db, 880005, is_pro=True)
    resp = client.get(
        f"/api/v1/internal/calc/{run.id}/export.xlsx",
        params={"telegram_user_id": 880005},
        headers=_headers(),
    )
    assert resp.status_code == 404, resp.text


def test_unknown_run_404(client, db):
    _make_user(db, 880006, is_pro=True)
    resp = client.get(
        "/api/v1/internal/calc/999999/export.xlsx",
        params={"telegram_user_id": 880006},
        headers=_headers(),
    )
    assert resp.status_code == 404, resp.text


def test_free_probing_other_run_sees_404_not_403(client, db):
    # Ownership is checked before the PRO gate, so a FREE user probing someone
    # else's run id leaks nothing beyond "not found".
    owner = _make_user(db, 880007, is_pro=True)
    run = _make_run(db, owner.id)
    _make_user(db, 880008, is_pro=False)
    resp = client.get(
        f"/api/v1/internal/calc/{run.id}/export.xlsx",
        params={"telegram_user_id": 880008},
        headers=_headers(),
    )
    assert resp.status_code == 404, resp.text


def test_internal_calc_returns_run_id(client, db, monkeypatch):
    # The just-computed result must carry the persisted run_id so the bot can
    # offer export on the result screen.
    from app.services.market import service as market_service
    from app.services.market.provider import RawMarketData

    market_service.reset_cache()
    monkeypatch.setattr(
        market_service,
        "fetch_market_data",
        lambda: RawMarketData(
            price_usdt=Decimal("60000"),
            network_difficulty=Decimal("80000000000000"),
            block_reward_btc=Decimal("3.125"),
        ),
    )
    resp = client.post(
        "/api/v1/internal/calc",
        json={
            "telegram_user_id": 880009,
            "hashrate_ths": "100",
            "power_w": 3250,
            "quantity": 1,
            "power_price": "0.05",
            "currency": "USDT",
        },
        headers=_headers(),
    )
    market_service.reset_cache()
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert isinstance(body["run_id"], int)
    assert body["run_id"] > 0
