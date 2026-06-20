"""Excel (.xlsx) export of a saved profitability calculation (Queue 2.2).

Pure, dependency-light report builder: it takes a plain :class:`CalcExportData`
snapshot (assembled from a ``calculation_runs`` row) and returns the workbook
bytes. No FastAPI, no aiogram, no DB session — so the generation can be unit
tested directly and reused by either the API endpoint or (in principle) the bot.

Money is written as real numbers with an Excel ``number_format`` (not strings),
so a user can sum/sort the cells in their spreadsheet. Decimals are converted to
float only at the cell-write boundary; all arithmetic upstream stays Decimal.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SITE_URL = "kirkalab.ru"
REPORT_TITLE = "Kirkalab — расчёт доходности"

# Currency display symbols (mirrors bot.history_format so the file reads the same
# as the on-screen card).
_CURRENCY_SYMBOL = {
    "USDT": "USDT",
    "USD": "$",
    "RUB": "₽",
    "CNY": "¥",
    "EUR": "€",
    "KZT": "₸",
}

# Excel number formats. Money keeps two decimals + a thin space thousands group;
# the per-kWh price keeps four decimals (matches the NUMERIC(12,4) storage).
_FMT_MONEY = '# ##0.00'
_FMT_PRICE = '# ##0.0000'
_FMT_HASHRATE = '# ##0.00'
_FMT_INT = '# ##0'

_TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
_SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


@dataclass(frozen=True)
class CalcExportData:
    """Flat snapshot of one calculation, everything the report needs.

    Assembled from a ``calculation_runs`` row (see
    :func:`app.api.v1.internal` export endpoint). Optional fields render as "—"
    or are simply omitted when absent, so a manual-entry calc (no model id) and a
    catalog calc both export cleanly."""

    run_id: int
    device_name: str
    quantity: int
    currency: str
    created_at: datetime | None
    hashrate_ths: Decimal | None = None
    power_w: int | None = None
    power_price: Decimal | None = None
    net_profit_day: Decimal | None = None
    net_profit_month: Decimal | None = None
    exchange_rate: Decimal | None = None


def _sym(currency: str) -> str:
    return _CURRENCY_SYMBOL.get((currency or "USDT").upper(), currency or "USDT")


def _num(value: Decimal | int | float | None) -> float | None:
    if value is None:
        return None
    return float(value)


def export_filename(run_id: int, created_at: datetime | None = None) -> str:
    """Stable, human-readable file name: ``kirkalab_calc_<id>_<YYYY-MM-DD>.xlsx``.

    Uses the calc date when available, otherwise today (UTC), so the name is
    always well-formed even for legacy rows without a timestamp."""
    moment = created_at or datetime.now(timezone.utc)
    return f"kirkalab_calc_{run_id}_{moment:%Y-%m-%d}.xlsx"


def build_calc_workbook(data: CalcExportData) -> bytes:
    """Render the calculation into a styled single-sheet workbook, return bytes.

    Layout: a title band, a "Параметры" block (model, qty, specs, price, rate),
    a "Результаты" block (daily/monthly net profit), and a footer with the site
    link. Money cells carry a numeric ``number_format`` so they stay numbers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Расчёт"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10

    currency = data.currency or "USDT"
    sym = _sym(currency)
    row = 1

    # --- Title band -------------------------------------------------------- #
    cell = ws.cell(row=row, column=1, value=REPORT_TITLE)
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = _TITLE_FILL
    cell.alignment = Alignment(vertical="center")
    ws.cell(row=row, column=2).fill = _TITLE_FILL
    ws.cell(row=row, column=3).fill = _TITLE_FILL
    ws.row_dimensions[row].height = 26
    row += 1

    when = data.created_at
    when_text = when.strftime("%d.%m.%Y %H:%M") if when else "—"
    ws.cell(row=row, column=1, value=f"Дата расчёта: {when_text}").font = Font(
        italic=True, color="595959"
    )
    row += 2

    # --- Параметры --------------------------------------------------------- #
    row = _section_header(ws, row, "Параметры")
    row = _kv(ws, row, "Модель ASIC", data.device_name or "—")
    row = _kv(ws, row, "Количество, шт.", _num(data.quantity), _FMT_INT)
    if data.hashrate_ths is not None:
        row = _kv(
            ws, row, "Хешрейт, TH/s", _num(data.hashrate_ths), _FMT_HASHRATE
        )
    if data.power_w is not None:
        row = _kv(ws, row, "Потребление, Вт", _num(data.power_w), _FMT_INT)
    if data.power_price is not None:
        row = _kv(
            ws,
            row,
            f"Цена э/э, {sym}/кВт·ч",
            _num(data.power_price),
            _FMT_PRICE,
        )
    row = _kv(ws, row, "Валюта", currency)
    if data.exchange_rate is not None:
        row = _kv(
            ws,
            row,
            f"Курс, {sym} за USDT",
            _num(data.exchange_rate),
            _FMT_PRICE,
        )
    row += 1

    # --- Результаты -------------------------------------------------------- #
    row = _section_header(ws, row, "Результаты")
    row = _kv(
        ws,
        row,
        f"Чистая прибыль в день, {sym}",
        _num(data.net_profit_day),
        _FMT_MONEY,
        bold=True,
    )
    row = _kv(
        ws,
        row,
        f"Чистая прибыль в месяц, {sym}",
        _num(data.net_profit_month),
        _FMT_MONEY,
        bold=True,
    )
    row += 1

    # --- Footer ------------------------------------------------------------ #
    footer = ws.cell(
        row=row, column=1, value=f"Сформировано в Kirkalab · {SITE_URL}"
    )
    footer.font = Font(italic=True, color="808080")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _section_header(ws, row: int, title: str) -> int:
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=12, color="1F4E78")
    cell.fill = _SECTION_FILL
    ws.cell(row=row, column=2).fill = _SECTION_FILL
    ws.cell(row=row, column=3).fill = _SECTION_FILL
    return row + 1


def _kv(
    ws,
    row: int,
    label: str,
    value,
    number_format: str | None = None,
    *,
    bold: bool = False,
) -> int:
    """Write one label/value pair; numeric values keep ``number_format``."""
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = Font(bold=bold)
    label_cell.border = _BORDER

    value_cell = ws.cell(row=row, column=2)
    if value is None:
        value_cell.value = "—"
    else:
        value_cell.value = value
        if number_format is not None:
            value_cell.number_format = number_format
    value_cell.font = Font(bold=bold)
    value_cell.alignment = Alignment(horizontal="right")
    value_cell.border = _BORDER
    return row + 1


__all__ = [
    "CalcExportData",
    "build_calc_workbook",
    "export_filename",
    "REPORT_TITLE",
    "SITE_URL",
]
